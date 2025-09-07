import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import cv2
import numpy as np
import matplotlib.pyplot as plt
import os
from PIL import Image, ImageTk
import threading
import colorsys

class 桥梁颜色分析器:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("桥梁颜色分析器 - 最终版")
        self.root.geometry("1400x900")
        
        # 设置英文字体
        plt.rcParams['font.family'] = 'DejaVu Sans'
        plt.rcParams['axes.unicode_minus'] = False
        
        # 数据存储
        self.image_paths = []
        self.images = []
        self.bridge_masks = []
        self.hsv_data = []
        self.color_analysis_results = []
        self.current_image_index = 0
        
        self.setup_ui()
        
    def setup_ui(self):
        """创建用户界面"""
        # 主框架
        main_frame = tk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 顶部控制区
        control_frame = tk.Frame(main_frame, bd=2, relief=tk.RAISED)
        control_frame.pack(fill=tk.X, pady=5)
        
        # 按钮区域
        button_frame = tk.Frame(control_frame)
        button_frame.pack(side=tk.LEFT, padx=5, pady=5)
        
        # 按钮
        tk.Button(button_frame, text="选择图像", command=self.select_images, 
                 font=('Arial', 10), bg='#4CAF50', fg='white', padx=15).pack(side=tk.LEFT, padx=2)
        
        tk.Button(button_frame, text="选择桥图文件夹", command=self.select_folder, 
                 font=('Arial', 10), bg='#4CAF50', fg='white', padx=15).pack(side=tk.LEFT, padx=2)
        
        tk.Button(button_frame, text="分析桥梁颜色", command=self.analyze_bridge_colors, 
                 font=('Arial', 10), bg='#2196F3', fg='white', padx=15).pack(side=tk.LEFT, padx=2)
        
        tk.Button(button_frame, text="显示HSV直方图", command=self.show_hsv_histograms, 
                 font=('Arial', 10), bg='#FF9800', fg='white', padx=15).pack(side=tk.LEFT, padx=2)
        
        tk.Button(button_frame, text="显示3D散点图", command=self.show_3d_scatter, 
                 font=('Arial', 10), bg='#9C27B0', fg='white', padx=15).pack(side=tk.LEFT, padx=2)
        
        tk.Button(button_frame, text="按段显示直方图", command=self.show_segment_histograms, 
                 font=('Arial', 10), bg='#00BCD4', fg='white', padx=15).pack(side=tk.LEFT, padx=2)
        
        tk.Button(button_frame, text="清除数据", command=self.clear_data, 
                 font=('Arial', 10), bg='#F44336', fg='white', padx=15).pack(side=tk.LEFT, padx=2)
        
        
        tk.Button(button_frame, text="导出HSV数据", command=self.export_hsv_data, 
                 font=('Arial', 10), bg='#E91E63', fg='white', padx=15).pack(side=tk.LEFT, padx=2)
        
        # 进度条和状态
        status_frame = tk.Frame(control_frame)
        status_frame.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=5, pady=5)
        
        self.progress_bar = ttk.Progressbar(status_frame, orient="horizontal", length=200, mode="determinate")
        self.progress_bar.pack(side=tk.LEFT, padx=5)
        
        self.status_label = tk.Label(status_frame, text="状态: 等待选择图像...", bd=1, relief=tk.SUNKEN, anchor=tk.W, font=('Arial', 9))
        self.status_label.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        # 主要内容区域
        content_frame = tk.Frame(main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # 左侧：图像显示区
        left_frame = tk.Frame(content_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        # 图像显示区
        image_display_frame = tk.Frame(left_frame, bd=2, relief=tk.SUNKEN)
        image_display_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # 原始图像
        tk.Label(image_display_frame, text="原始图像", font=('Arial', 12, 'bold')).pack(pady=5)
        self.original_image_label = tk.Label(image_display_frame, bg='white', width=400, height=300, text="等待选择图像...", font=('Arial', 16), fg='gray')
        self.original_image_label.pack(padx=5, pady=5)
        
        # 桥梁掩码
        tk.Label(image_display_frame, text="桥梁掩码", font=('Arial', 12, 'bold')).pack(pady=5)
        self.mask_image_label = tk.Label(image_display_frame, bg='white', width=400, height=300, text="等待分析...", font=('Arial', 16), fg='gray')
        self.mask_image_label.pack(padx=5, pady=5)
        
        # 右侧：颜色分析结果区
        right_frame = tk.Frame(content_frame, bd=2, relief=tk.SUNKEN)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, padx=(5, 0))
        
        # 颜色分析结果标题
        tk.Label(right_frame, text="桥梁颜色分析结果", font=('Arial', 14, 'bold')).pack(pady=5)
        
        # 创建颜色分析结果显示区域
        self.create_color_analysis_panel(right_frame)
        
        # 导航按钮 - 确保在底部显示
        nav_frame = tk.Frame(main_frame, bd=2, relief=tk.RAISED)
        nav_frame.pack(fill=tk.X, pady=10)
        
        # 导航按钮居中显示
        nav_center_frame = tk.Frame(nav_frame)
        nav_center_frame.pack(expand=True)
        
        tk.Button(nav_center_frame, text="上一张", command=self.show_previous_image, 
                 font=('Arial', 12), bg='#607D8B', fg='white', padx=20, pady=5).pack(side=tk.LEFT, padx=10)
        
        self.image_count_label = tk.Label(nav_center_frame, text="0/0", font=('Arial', 14, 'bold'))
        self.image_count_label.pack(side=tk.LEFT, padx=30)
        
        tk.Button(nav_center_frame, text="下一张", command=self.show_next_image, 
                 font=('Arial', 12), bg='#607D8B', fg='white', padx=20, pady=5).pack(side=tk.LEFT, padx=10)
        
    def create_color_analysis_panel(self, parent):
        """创建颜色分析结果显示面板"""
        # 创建滚动框架
        canvas = tk.Canvas(parent, width=350, height=600, bg='white')
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 颜色分析结果列表
        self.color_results_frame = scrollable_frame
        
        # 添加初始提示文本
        tk.Label(scrollable_frame, text="等待颜色分析结果...", 
                font=('Arial', 12), fg='gray', bg='white').pack(pady=20)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
    def select_folder(self):
        """选择桥图文件夹"""
        folder_path = filedialog.askdirectory(title="选择桥图文件夹")
        
        if folder_path:
            image_files = []
            for root, _, files in os.walk(folder_path):
                for file in files:
                    if file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif')):
                        image_files.append(os.path.join(root, file))
            
            if image_files:
                self.image_paths = image_files
                self.status_label.config(text=f"状态: 已从文件夹中选择 {len(self.image_paths)} 张图像。")
                self.progress_bar['maximum'] = len(self.image_paths)
                self.progress_bar['value'] = 0
                self.update_image_count_label()
                self.current_image_index = 0
                self.display_current_image()
                messagebox.showinfo("成功", f"已选择桥图文件夹\n共找到 {len(self.image_paths)} 张图像")
            else:
                messagebox.showinfo("提示", "所选文件夹中没有找到支持的图像文件。")
        else:
            self.status_label.config(text="未选择文件夹")
    
    def select_images(self):
        """选择图像文件"""
        file_paths = filedialog.askopenfilenames(
            title="选择桥梁图像",
            filetypes=[("图像文件", "*.jpg *.jpeg *.png *.bmp *.tiff")]
        )
        
        if file_paths:
            self.image_paths = list(file_paths)
            self.status_label.config(text=f"状态: 已选择 {len(self.image_paths)} 张图像。")
            self.progress_bar['maximum'] = len(self.image_paths)
            self.progress_bar['value'] = 0
            self.update_image_count_label()
            self.current_image_index = 0
            self.display_current_image()
            messagebox.showinfo("成功", f"已选择 {len(self.image_paths)} 张图像")
        else:
            self.status_label.config(text="未选择图像")
    
    def analyze_bridge_colors(self):
        """分析桥梁颜色"""
        if not self.image_paths:
            messagebox.showwarning("警告", "请先选择图像文件！")
            return
        
        # 在新线程中处理，避免界面冻结
        thread = threading.Thread(target=self._process_images)
        thread.daemon = True
        thread.start()
    
    def _process_images(self):
        """处理图像（在后台线程中运行）"""
        self.images = []
        self.bridge_masks = []
        self.hsv_data = []
        self.color_analysis_results = []
        
        total_images = len(self.image_paths)
        processed_count = 0
        failed_images = []  # 记录失败的图像
        
        for i, image_path in enumerate(self.image_paths):
            try:
                # 更新进度
                self.root.after(0, lambda p=i+1, t=total_images: 
                              self.progress_bar.configure(value=p))
                
                # 更新状态
                self.root.after(0, lambda p=i+1, t=total_images: 
                              self.status_label.config(text=f"状态: 正在处理图像 {p}/{t}: {os.path.basename(image_path)}"))
                
                # 读取图像
                image = self.read_image_safely(image_path)
                if image is None:
                    failed_images.append(f"{os.path.basename(image_path)}: 图像读取失败")
                    continue
                
                self.images.append(image)
                
                # 桥梁分割
                bridge_mask = self.segment_bridge(image)
                
                if bridge_mask is None:
                    failed_images.append(f"{os.path.basename(image_path)}: 桥梁分割失败")
                    continue
                
                self.bridge_masks.append(bridge_mask)
                
                # 提取HSV数据
                hsv_data = self.extract_hsv_data(image, bridge_mask)
                if hsv_data is None or len(hsv_data) == 0:
                    failed_images.append(f"{os.path.basename(image_path)}: HSV提取失败")
                    continue
                
                self.hsv_data.append(hsv_data)
                
                # 分析颜色
                color_result = self.analyze_colors(hsv_data, os.path.basename(image_path))
                if color_result is None:
                    failed_images.append(f"{os.path.basename(image_path)}: 颜色分析失败")
                    continue
                
                self.color_analysis_results.append(color_result)
                processed_count += 1
                
                # 显示第一张图像的处理结果
                if i == 0:
                    self.root.after(0, self.display_current_image)
                    
            except Exception as e:
                failed_images.append(f"{os.path.basename(image_path)}: 处理异常 - {str(e)}")
                print(f"处理图像 {image_path} 时出错: {e}")
                continue
        
        # 处理完成
        self.root.after(0, lambda: self.progress_bar.configure(value=total_images))
        
        # 显示详细的处理结果
        if failed_images:
            status_text = f"状态: 分析完成, 共处理 {processed_count}/{total_images} 张图像。失败 {len(failed_images)} 张。"
            self.root.after(0, lambda: self.status_label.config(text=status_text))
            self.root.after(0, lambda: messagebox.showinfo("完成", 
                f"桥梁颜色分析完成！\n\n成功处理: {processed_count} 张\n失败: {len(failed_images)} 张\n总计: {total_images} 张\n\n失败详情:\n" + "\n".join(failed_images[:10])))
        else:
            self.root.after(0, lambda: self.status_label.config(text=f"状态: 分析完成, 共处理 {processed_count}/{total_images} 张图像。"))
            self.root.after(0, lambda: messagebox.showinfo("完成", f"桥梁颜色分析完成！\n共处理 {processed_count}/{total_images} 张图像"))
        
        self.root.after(0, self.display_current_image)
        self.root.after(0, self.update_color_analysis_display)
    
    def read_image_safely(self, image_path):
        """安全读取图像，处理各种格式和错误"""
        try:
            # 尝试使用PIL读取
            pil_image = Image.open(image_path)
            image = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)
            return image
        except Exception as e:
            print(f"PIL读取失败，尝试OpenCV: {e}")
            try:
                # 备用方案：使用OpenCV
                image = cv2.imdecode(np.fromfile(image_path, dtype=np.uint8), cv2.IMREAD_COLOR)
                return image
            except Exception as e2:
                print(f"OpenCV读取也失败: {e2}")
                return None
    
    def analyze_colors(self, hsv_data, image_name):
        """分析HSV数据，提取主要颜色、次要颜色和点缀颜色"""
        if hsv_data is None or len(hsv_data) == 0:
            return None
            
        # 找到HSV值最多的区域（主要颜色）
        h_bins = np.histogram(hsv_data[:, 0], bins=18, range=(0, 180))[0]
        s_bins = np.histogram(hsv_data[:, 1], bins=25, range=(0, 255))[0]
        v_bins = np.histogram(hsv_data[:, 2], bins=25, range=(0, 255))[0]
        
        # 主要颜色：HSV值最多的区域
        primary_h = np.argmax(h_bins) * 10
        primary_s = np.argmax(s_bins) * 10
        primary_v = np.argmax(v_bins) * 10
        
        # 次要颜色：第二多的区域
        h_bins_sorted = np.argsort(h_bins)[::-1]
        s_bins_sorted = np.argsort(s_bins)[::-1]
        v_bins_sorted = np.argsort(v_bins)[::-1]
        
        secondary_h = h_bins_sorted[1] * 10 if len(h_bins_sorted) > 1 else primary_h
        secondary_s = s_bins_sorted[1] * 10 if len(s_bins_sorted) > 1 else primary_s
        secondary_v = v_bins_sorted[1] * 10 if len(v_bins_sorted) > 1 else primary_v
        
        # 点缀颜色：随机选择一些有代表性的颜色
        if len(hsv_data) > 100:
            sample_indices = np.random.choice(len(hsv_data), 100, replace=False)
            sample_data = hsv_data[sample_indices]
        else:
            sample_data = hsv_data
            
        spot_colors = []
        for _ in range(3):
            if len(sample_data) > 0:
                idx = np.random.randint(0, len(sample_data))
                spot_colors.append(sample_data[idx])
        
        # 转换为Munsell颜色表示
        primary_munsell = self.hsv_to_munsell(primary_h, primary_s, primary_v)
        secondary_munsell = self.hsv_to_munsell(secondary_h, secondary_s, secondary_v)
        spot_munsells = [self.hsv_to_munsell(h, s, v) for h, s, v in spot_colors]
        
        return {
            'image_name': image_name,
            'primary_color': {
                'hsv': (primary_h, primary_s, primary_v),
                'munsell': primary_munsell,
                'rgb': self.hsv_to_rgb(primary_h, primary_s, primary_v)
            },
            'secondary_color': {
                'hsv': (secondary_h, secondary_s, secondary_v),
                'munsell': secondary_munsell,
                'rgb': self.hsv_to_rgb(secondary_h, secondary_s, secondary_v)
            },
            'spot_colors': [{
                'hsv': (h, s, v),
                'munsell': munsell,
                'rgb': self.hsv_to_rgb(h, s, v)
            } for (h, s, v), munsell in zip(spot_colors, spot_munsells)]
        }
    
    def get_weighted_hue(self, hsv_data, h_bins):
        """获取加权平均色调值"""
        try:
            # 找到峰值区域
            peak_indices = np.argsort(h_bins)[-3:]  # 前3个峰值
            
            total_weight = 0
            weighted_sum = 0
            
            for idx in peak_indices:
                if h_bins[idx] > 0:
                    weight = h_bins[idx]
                    hue_value = idx * 5  # 5度间隔
                    weighted_sum += hue_value * weight
                    total_weight += weight
            
            if total_weight > 0:
                return int(weighted_sum / total_weight)
            else:
                return 90  # 默认值
        except:
            return 90
    
    def get_weighted_saturation(self, hsv_data, s_bins):
        """获取加权平均饱和度值"""
        try:
            peak_indices = np.argsort(s_bins)[-3:]
            
            total_weight = 0
            weighted_sum = 0
            
            for idx in peak_indices:
                if s_bins[idx] > 0:
                    weight = s_bins[idx]
                    sat_value = idx * 5  # 5间隔
                    weighted_sum += sat_value * weight
                    total_weight += weight
            
            if total_weight > 0:
                return int(weighted_sum / total_weight)
            else:
                return 128
        except:
            return 128
    
    def get_weighted_value(self, hsv_data, v_bins):
        """获取加权平均明度值"""
        try:
            peak_indices = np.argsort(v_bins)[-3:]
            
            total_weight = 0
            weighted_sum = 0
            
            for idx in peak_indices:
                if v_bins[idx] > 0:
                    weight = v_bins[idx]
                    val_value = idx * 5
                    weighted_sum += val_value * weight
                    total_weight += weight
            
            if total_weight > 0:
                return int(weighted_sum / total_weight)
            else:
                return 128
        except:
            return 128
    
    def hsv_to_munsell(self, h, s, v):
        """将HSV值转换为Munsell颜色表示"""
        if v < 30:
            return f"N {int(v/10)}"
        elif s < 20:
            return f"N {int(v/10)}"
        else:
            # 根据H值确定色相
            if h < 15 or h >= 165:
                hue = "R"
            elif h < 45:
                hue = "YR"
            elif h < 75:
                hue = "Y"
            elif h < 105:
                hue = "GY"
            elif h < 135:
                hue = "G"
            elif h < 165:
                hue = "BG"
            else:
                hue = "B"
            
            # 色相值（1-10）
            hue_value = int((h % 30) / 3) + 1
            # 明度值（1-10）
            value = int(v / 25.5) + 1
            # 彩度值（1-10）
            chroma = int(s / 25.5) + 1
            
            return f"{hue} {value}/{chroma}"
    
    def hsv_to_rgb(self, h, s, v):
        """将HSV值转换为RGB值"""
        h = h / 180.0
        s = s / 255.0
        v = v / 255.0
        
        rgb = colorsys.hsv_to_rgb(h, s, v)
        return tuple(int(x * 255) for x in rgb)
    
    def update_color_analysis_display(self):
        """更新颜色分析结果显示"""
        # 清除之前的内容
        for widget in self.color_results_frame.winfo_children():
            widget.destroy()
        
        if not self.color_analysis_results:
            tk.Label(self.color_results_frame, text="暂无颜色分析结果", 
                    font=('Arial', 12), fg='gray').pack(pady=20)
            return
        
        # 显示每张图像的颜色分析结果
        for i, result in enumerate(self.color_analysis_results):
            if result is None:
                continue
                
            # 创建图像结果框架
            result_frame = tk.Frame(self.color_results_frame, bd=2, relief=tk.RAISED)
            result_frame.pack(fill=tk.X, padx=5, pady=5)
            
            # 图像名称
            tk.Label(result_frame, text=f"图像 {i+1}: {result['image_name']}", 
                    font=('Arial', 10, 'bold')).pack(anchor=tk.W, padx=5, pady=2)
            
            # 主要颜色
            primary_frame = tk.Frame(result_frame)
            primary_frame.pack(fill=tk.X, padx=5, pady=2)
            
            tk.Label(primary_frame, text="主要颜色:", font=('Arial', 9, 'bold')).pack(side=tk.LEFT)
            
            # 颜色显示框
            primary_color_frame = tk.Frame(primary_frame, width=30, height=20, 
                                         bg=self.rgb_to_hex(result['primary_color']['rgb']))
            primary_color_frame.pack(side=tk.LEFT, padx=5)
            primary_color_frame.pack_propagate(False)
            
            tk.Label(primary_frame, text=result['primary_color']['munsell'], 
                    font=('Arial', 8)).pack(side=tk.LEFT, padx=5)
            
            # 次要颜色
            secondary_frame = tk.Frame(result_frame)
            secondary_frame.pack(fill=tk.X, padx=5, pady=2)
            
            tk.Label(secondary_frame, text="次要颜色:", font=('Arial', 9, 'bold')).pack(side=tk.LEFT)
            
            secondary_color_frame = tk.Frame(secondary_frame, width=30, height=20, 
                                          bg=self.rgb_to_hex(result['secondary_color']['rgb']))
            secondary_color_frame.pack(side=tk.LEFT, padx=5)
            secondary_color_frame.pack_propagate(False)
            
            tk.Label(secondary_frame, text=result['secondary_color']['munsell'], 
                    font=('Arial', 8)).pack(side=tk.LEFT, padx=5)
            
            # 点缀颜色
            spot_frame = tk.Frame(result_frame)
            spot_frame.pack(fill=tk.X, padx=5, pady=2)
            
            tk.Label(spot_frame, text="点缀颜色:", font=('Arial', 9, 'bold')).pack(side=tk.LEFT)
            
            for j, spot_color in enumerate(result['spot_colors'][:3]):
                spot_color_frame = tk.Frame(spot_frame, width=20, height=15, 
                                         bg=self.rgb_to_hex(spot_color['rgb']))
                spot_color_frame.pack(side=tk.LEFT, padx=2)
                spot_color_frame.pack_propagate(False)
                
                tk.Label(spot_frame, text=spot_color['munsell'], 
                        font=('Arial', 7)).pack(side=tk.LEFT, padx=2)
            
            # 分隔线
            tk.Frame(result_frame, height=1, bg='gray').pack(fill=tk.X, pady=2)
    
    def rgb_to_hex(self, rgb):
        """将RGB元组转换为十六进制颜色字符串"""
        return f'#{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}'
    
    def segment_bridge(self, image):
        """简单有效的桥梁分割（稳定版本）"""
        try:
            # 转换为HSV颜色空间
            hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
            
            # 1. 基础桥梁颜色检测（保守范围）
            # 石桥/混凝土桥：浅灰到白色
            lower_bridge = np.array([0, 0, 100])      # 提高明度下限，避免天空
            upper_bridge = np.array([180, 60, 255])   # 降低饱和度上限
            
            # 创建基础掩码
            bridge_mask = cv2.inRange(hsv, lower_bridge, upper_bridge)
            
            # 2. 天空过滤（只过滤明显的蓝色天空）
            # 天空：高亮度、低饱和度的蓝色
            lower_sky = np.array([100, 10, 200])      # 更严格的天空条件
            upper_sky = np.array([140, 60, 255])
            sky_mask = cv2.inRange(hsv, lower_sky, upper_sky)
            
            # 3. 从桥梁掩码中减去天空
            bridge_mask = cv2.bitwise_and(bridge_mask, cv2.bitwise_not(sky_mask))
            
            # 4. 形态学清理（简单有效）
            # 先开运算去除小噪点
            kernel_open = np.ones((5, 5), np.uint8)
            bridge_mask = cv2.morphologyEx(bridge_mask, cv2.MORPH_OPEN, kernel_open)
            
            # 再闭运算填充小孔洞
            kernel_close = np.ones((9, 9), np.uint8)
            bridge_mask = cv2.morphologyEx(bridge_mask, cv2.MORPH_CLOSE, kernel_close)
            
            # 5. 轮廓筛选（简单规则）
            contours, _ = cv2.findContours(bridge_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            if contours:
                # 找到最大的轮廓
                largest_contour = max(contours, key=cv2.contourArea)
                area = cv2.contourArea(largest_contour)
                
                # 面积过滤：不能太小，也不能太大
                h, w = image.shape[:2]
                min_area = h * w * 0.01   # 至少1%
                max_area = h * w * 0.7    # 最多70%
                
                if min_area < area < max_area:
                    # 创建最终掩码
                    final_mask = np.zeros_like(bridge_mask)
                    cv2.fillPoly(final_mask, [largest_contour], 255)
                    
                    # 轻微膨胀平滑边缘
                    kernel = np.ones((3, 3), np.uint8)
                    final_mask = cv2.dilate(final_mask, kernel, iterations=1)
                    
                    return final_mask
            
            # 6. 如果上述方法失败，使用简单的备用方案
            return self.simple_fallback_detection(image)
            
        except Exception as e:
            print(f"简单桥梁分割出错: {e}")
            return self.simple_fallback_detection(image)
    
    def multi_color_analysis(self, image):
        """增强版多颜色空间分析（带天空和植被过滤）"""
        try:
            # HSV颜色空间
            hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
            
            # 1. 桥梁材质颜色范围（更精确）
            # 石桥：浅灰到白色
            lower_stone = np.array([0, 0, 90])      # 提高明度下限，避免天空
            upper_stone = np.array([180, 70, 255])  # 降低饱和度上限
            
            # 混凝土桥：灰色到米色
            lower_concrete = np.array([0, 0, 70])
            upper_concrete = np.array([180, 90, 210])
            
            # 钢桥：深灰到银色
            lower_steel = np.array([0, 0, 50])
            upper_steel = np.array([180, 70, 190])
            
            # 创建多材质掩码
            stone_mask = cv2.inRange(hsv, lower_stone, upper_stone)
            concrete_mask = cv2.inRange(hsv, lower_concrete, upper_concrete)
            steel_mask = cv2.inRange(hsv, lower_steel, upper_steel)
            
            # 合并掩码
            bridge_color_mask = cv2.bitwise_or(stone_mask, concrete_mask)
            bridge_color_mask = cv2.bitwise_or(bridge_color_mask, steel_mask)
            
            # 2. 天空过滤（排除亮蓝色区域）
            # 天空通常是高亮度、低饱和度的蓝色
            lower_sky = np.array([100, 20, 180])  # 蓝色，低饱和度，高亮度
            upper_sky = np.array([130, 80, 255])
            sky_mask = cv2.inRange(hsv, lower_sky, upper_sky)
            
            # 3. 植被过滤（排除绿色区域）
            # 植被通常是中等亮度、高饱和度的绿色
            lower_vegetation = np.array([35, 80, 40])   # 绿色，高饱和度
            upper_vegetation = np.array([85, 255, 200])
            vegetation_mask = cv2.inRange(hsv, lower_vegetation, upper_vegetation)
            
            # 4. 水过滤（排除深色区域）
            # 水通常是低亮度、低饱和度的深色
            lower_water = np.array([0, 0, 0])
            upper_water = np.array([180, 100, 80])
            water_mask = cv2.inRange(hsv, lower_water, upper_water)
            
            # 5. 组合过滤：桥梁颜色 - (天空 + 植被 + 水)
            background_mask = cv2.bitwise_or(sky_mask, vegetation_mask)
            background_mask = cv2.bitwise_or(background_mask, water_mask)
            
            # 从桥梁掩码中减去背景
            final_color_mask = cv2.bitwise_and(bridge_color_mask, cv2.bitwise_not(background_mask))
            
            # 6. 形态学清理和优化
            # 先开运算去除小噪点
            kernel_open = np.ones((3, 3), np.uint8)
            final_color_mask = cv2.morphologyEx(final_color_mask, cv2.MORPH_OPEN, kernel_open)
            
            # 再闭运算填充小孔洞
            kernel_close = np.ones((7, 7), np.uint8)
            final_color_mask = cv2.morphologyEx(final_color_mask, cv2.MORPH_CLOSE, kernel_close)
            
            # 最后开运算平滑边缘
            final_color_mask = cv2.morphologyEx(final_color_mask, cv2.MORPH_OPEN, kernel_open)
            
            return final_color_mask
            
        except Exception as e:
            print(f"增强版多颜色分析失败: {e}")
            # 返回简单的灰度掩码
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            _, mask = cv2.threshold(gray, 120, 255, cv2.THRESH_BINARY)
            return mask
    
    def texture_analysis(self, image):
        """增强版纹理特征分析（专门针对桥梁结构）"""
        try:
            # 转换为灰度图
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            # 1. 局部二值模式（LBP）纹理分析
            # 桥梁通常有规则的纹理模式
            lbp_mask = self.compute_lbp_texture(gray)
            
            # 2. 梯度纹理分析
            gradient_mask = self.compute_gradient_texture(gray)
            
            # 3. 结构张量分析
            structure_mask = self.compute_structure_tensor(gray)
            
            # 4. 融合多种纹理特征
            texture_mask = cv2.addWeighted(lbp_mask, 0.4, gradient_mask, 0.4, 0)
            texture_mask = cv2.addWeighted(texture_mask, 0.8, structure_mask, 0.2, 0)
            
            # 5. 二值化和清理
            _, texture_mask = cv2.threshold(texture_mask, 127, 255, cv2.THRESH_BINARY)
            
            # 形态学操作清理
            kernel = np.ones((3, 3), np.uint8)
            texture_mask = cv2.morphologyEx(texture_mask, cv2.MORPH_OPEN, kernel)
            
            return texture_mask
            
        except Exception as e:
            print(f"增强版纹理分析失败: {e}")
            return np.zeros(image.shape[:2], dtype=np.uint8)
    
    def compute_lbp_texture(self, gray):
        """计算局部二值模式纹理"""
        try:
            # 简化的LBP实现
            height, width = gray.shape
            lbp = np.zeros_like(gray)
            
            for i in range(1, height-1):
                for j in range(1, width-1):
                    center = gray[i, j]
                    code = 0
                    
                    # 8邻域LBP
                    neighbors = [
                        gray[i-1, j-1], gray[i-1, j], gray[i-1, j+1],
                        gray[i, j+1], gray[i+1, j+1], gray[i+1, j],
                        gray[i+1, j-1], gray[i, j-1]
                    ]
                    
                    for k, neighbor in enumerate(neighbors):
                        if neighbor >= center:
                            code |= (1 << k)
                    
                    lbp[i, j] = code
            
            # 归一化到0-255
            lbp = (lbp * 255 / 255).astype(np.uint8)
            
            # 阈值化
            _, lbp_mask = cv2.threshold(lbp, 127, 255, cv2.THRESH_BINARY)
            
            return lbp_mask
            
        except Exception as e:
            print(f"LBP纹理计算失败: {e}")
            return np.zeros(gray.shape, dtype=np.uint8)
    
    def compute_gradient_texture(self, gray):
        """计算梯度纹理"""
        try:
            # Sobel梯度
            grad_x = cv2.Sobel(gray, cv2.CV_64F, 1, 0, ksize=3)
            grad_y = cv2.Sobel(gray, cv2.CV_64F, 0, 1, ksize=3)
            
            # 梯度幅值
            gradient_magnitude = np.sqrt(grad_x**2 + grad_y**2)
            
            # 归一化到0-255
            gradient_magnitude = (gradient_magnitude * 255 / gradient_magnitude.max()).astype(np.uint8)
            
            # 阈值化
            _, gradient_mask = cv2.threshold(gradient_magnitude, 50, 255, cv2.THRESH_BINARY)
            
            return gradient_mask
            
        except Exception as e:
            print(f"梯度纹理计算失败: {e}")
            return np.zeros(gray.shape, dtype=np.uint8)
    
    def compute_structure_tensor(self, gray):
        """计算结构张量"""
        try:
            # 计算梯度
            grad_x = cv2.Sobel(gray, cv2.CV_64F, 1, 0, ksize=3)
            grad_y = cv2.Sobel(gray, cv2.CV_64F, 0, 1, ksize=3)
            
            # 结构张量元素
            Ixx = grad_x * grad_x
            Ixy = grad_x * grad_y
            Iyy = grad_y * grad_y
            
            # 高斯平滑
            Ixx = cv2.GaussianBlur(Ixx, (5, 5), 1)
            Ixy = cv2.GaussianBlur(Ixy, (5, 5), 1)
            Iyy = cv2.GaussianBlur(Iyy, (5, 5), 1)
            
            # 计算特征值
            trace = Ixx + Iyy
            det = Ixx * Iyy - Ixy * Ixy
            
            # 结构一致性（桥梁通常有高结构一致性）
            structure_consistency = det / (trace + 1e-6)
            
            # 归一化到0-255
            structure_consistency = (structure_consistency * 255 / (structure_consistency.max() + 1e-6)).astype(np.uint8)
            
            # 阈值化
            _, structure_mask = cv2.threshold(structure_consistency, 100, 255, cv2.THRESH_BINARY)
            
            return structure_mask
            
        except Exception as e:
            print(f"结构张量计算失败: {e}")
            return np.zeros(gray.shape, dtype=np.uint8)
    
    def edge_analysis(self, image):
        """增强版边缘特征分析（专门针对桥梁结构）"""
        try:
            # 转换为灰度图
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            # 1. 多尺度边缘检测
            edges_multi = self.multi_scale_edge_detection(gray)
            
            # 2. 直线检测和连接
            line_mask = self.enhanced_line_detection(gray, edges_multi)
            
            # 3. 轮廓边缘检测
            contour_mask = self.contour_edge_detection(gray)
            
            # 4. 融合多种边缘特征
            edge_mask = cv2.addWeighted(line_mask, 0.6, contour_mask, 0.4, 0)
            
            # 5. 形态学优化
            kernel = np.ones((3, 3), np.uint8)
            edge_mask = cv2.morphologyEx(edge_mask, cv2.MORPH_CLOSE, kernel)
            
            return edge_mask
            
        except Exception as e:
            print(f"增强版边缘分析失败: {e}")
            return np.zeros(image.shape[:2], dtype=np.uint8)
    
    def multi_scale_edge_detection(self, gray):
        """多尺度边缘检测"""
        try:
            # 多个高斯模糊核大小
            scales = [3, 5, 7]
            edges_combined = np.zeros_like(gray)
            
            for scale in scales:
                # 高斯模糊
                blurred = cv2.GaussianBlur(gray, (scale, scale), 0)
                
                # Canny边缘检测（自适应阈值）
                mean_val = np.mean(blurred)
                std_val = np.std(blurred)
                
                low_threshold = max(0, mean_val - 0.8 * std_val)
                high_threshold = min(255, mean_val + 0.8 * std_val)
                
                edges = cv2.Canny(blurred, low_threshold, high_threshold)
                
                # 累加边缘
                edges_combined = cv2.bitwise_or(edges_combined, edges)
            
            return edges_combined
            
        except Exception as e:
            print(f"多尺度边缘检测失败: {e}")
            return np.zeros(gray.shape, dtype=np.uint8)
    
    def enhanced_line_detection(self, gray, edges):
        """增强版直线检测"""
        try:
            # 霍夫线变换检测直线
            lines = cv2.HoughLinesP(edges, 1, np.pi/180, threshold=30, 
                                  minLineLength=20, maxLineGap=15)
            
            if lines is None:
                return np.zeros_like(gray)
            
            # 创建线掩码
            line_mask = np.zeros_like(gray)
            
            # 过滤和优化直线
            filtered_lines = []
            for line in lines:
                x1, y1, x2, y2 = line[0]
                
                # 计算线段长度和角度
                length = np.sqrt((x2-x1)**2 + (y2-y1)**2)
                angle = np.arctan2(y2-y1, x2-x1) * 180 / np.pi
                
                # 过滤条件：长度和角度
                if length > 15 and (abs(angle) < 30 or abs(angle) > 150 or 
                                   abs(angle - 90) < 30 or abs(angle - 270) < 30):
                    filtered_lines.append(line)
            
            # 绘制过滤后的直线
            for line in filtered_lines:
                x1, y1, x2, y2 = line[0]
                cv2.line(line_mask, (x1, y1), (x2, y2), 255, 3)
            
            # 膨胀线掩码
            kernel = np.ones((5, 5), np.uint8)
            line_mask = cv2.dilate(line_mask, kernel, iterations=2)
            
            return line_mask
            
        except Exception as e:
            print(f"增强版直线检测失败: {e}")
            return np.zeros(gray.shape, dtype=np.uint8)
    
    def contour_edge_detection(self, gray):
        """轮廓边缘检测"""
        try:
            # 使用Laplacian检测边缘
            laplacian = cv2.Laplacian(gray, cv2.CV_64F)
            laplacian = np.uint8(np.absolute(laplacian))
            
            # 阈值化
            _, contour_mask = cv2.threshold(laplacian, 50, 255, cv2.THRESH_BINARY)
            
            # 形态学清理
            kernel = np.ones((2, 2), np.uint8)
            contour_mask = cv2.morphologyEx(contour_mask, cv2.MORPH_OPEN, kernel)
            
            return contour_mask
            
        except Exception as e:
            print(f"轮廓边缘检测失败: {e}")
            return np.zeros(gray.shape, dtype=np.uint8)
    
    def feature_fusion(self, color_mask, texture_mask, edge_mask):
        """增强版多特征融合（自适应权重）"""
        try:
            # 1. 计算各掩码的质量分数
            color_score = self.calculate_mask_quality(color_mask)
            texture_score = self.calculate_mask_quality(texture_mask)
            edge_score = self.calculate_mask_quality(edge_mask)
            
            # 2. 自适应权重分配
            total_score = color_score + texture_score + edge_score
            if total_score > 0:
                color_weight = color_score / total_score
                texture_weight = texture_score / total_score
                edge_weight = edge_score / total_score
            else:
                # 默认权重
                color_weight, texture_weight, edge_weight = 0.5, 0.3, 0.2
            
            # 3. 加权融合
            fused_mask = cv2.addWeighted(color_mask, color_weight, texture_mask, texture_weight, 0)
            fused_mask = cv2.addWeighted(fused_mask, 0.8, edge_mask, edge_weight, 0)
            
            # 4. 自适应阈值化
            threshold = self.calculate_adaptive_threshold(fused_mask)
            _, fused_mask = cv2.threshold(fused_mask, threshold, 255, cv2.THRESH_BINARY)
            
            # 5. 智能形态学优化
            fused_mask = self.smart_morphology_optimization(fused_mask)
            
            return fused_mask
            
        except Exception as e:
            print(f"增强版特征融合失败: {e}")
            return color_mask  # 返回颜色掩码作为备用
    
    def calculate_mask_quality(self, mask):
        """计算掩码质量分数"""
        try:
            if mask is None or np.sum(mask) == 0:
                return 0
            
            # 计算连通区域数量（越少越好）
            contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            contour_count = len(contours)
            
            # 计算掩码的紧凑性
            total_pixels = np.sum(mask > 0)
            if total_pixels == 0:
                return 0
            
            # 计算边界像素比例（边界越少越好）
            kernel = np.ones((3, 3), np.uint8)
            eroded = cv2.erode(mask, kernel)
            boundary_pixels = total_pixels - np.sum(eroded > 0)
            boundary_ratio = boundary_pixels / total_pixels
            
            # 质量分数：像素数量 - 连通区域惩罚 - 边界惩罚
            quality_score = total_pixels - contour_count * 100 - boundary_ratio * 1000
            
            return max(0, quality_score)
            
        except Exception as e:
            print(f"掩码质量计算失败: {e}")
            return 0
    
    def calculate_adaptive_threshold(self, mask):
        """计算自适应阈值"""
        try:
            # 使用Otsu方法计算最佳阈值
            if mask.dtype != np.uint8:
                mask = (mask * 255 / mask.max()).astype(np.uint8)
            
            # 如果掩码已经是二值的，直接返回
            if len(np.unique(mask)) <= 2:
                return 127
            
            # 计算直方图
            hist = cv2.calcHist([mask], [0], None, [256], [0, 256])
            
            # 找到非零像素的最小值
            non_zero_indices = np.where(hist > 0)[0]
            if len(non_zero_indices) > 0:
                min_val = non_zero_indices[0]
                max_val = non_zero_indices[-1]
                
                # 自适应阈值：在非零范围内选择中间值
                threshold = min_val + (max_val - min_val) * 0.3
                return int(threshold)
            
            return 127
            
        except Exception as e:
            print(f"自适应阈值计算失败: {e}")
            return 127
    
    def smart_morphology_optimization(self, mask):
        """智能形态学优化"""
        try:
            # 1. 分析掩码结构
            contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            if not contours:
                return mask
            
            # 2. 计算掩码的几何特性
            total_area = np.sum(mask > 0)
            largest_contour = max(contours, key=cv2.contourArea)
            largest_area = cv2.contourArea(largest_contour)
            
            # 3. 根据结构选择形态学操作
            if largest_area / total_area > 0.8:
                # 主要是一个大区域，使用温和的形态学操作
                kernel = np.ones((5, 5), np.uint8)
                mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
                mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
            else:
                # 多个小区域，使用更强的形态学操作
                kernel = np.ones((7, 7), np.uint8)
                mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
                mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
                
                # 连接相近的区域
                kernel = np.ones((3, 3), np.uint8)
                mask = cv2.dilate(mask, kernel, iterations=1)
            
            return mask
            
        except Exception as e:
            print(f"智能形态学优化失败: {e}")
            return mask
    
    def intelligent_contour_selection(self, image, mask):
        """智能轮廓筛选"""
        try:
            # 查找轮廓
            contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            if not contours:
                return None
            
            h, w = image.shape[:2]
            image_area = h * w
            
            # 自适应面积阈值
            min_area = image_area * 0.005  # 0.5%
            max_area = image_area * 0.8    # 80%
            
            # 筛选有效轮廓
            valid_contours = []
            for contour in contours:
                area = cv2.contourArea(contour)
                if min_area < area < max_area:
                    # 计算轮廓的几何特征
                    x, y, w_rect, h_rect = cv2.boundingRect(contour)
                    aspect_ratio = w_rect / h_rect
                    
                    # 更灵活的几何条件
                    # 桥梁可以是横向或纵向
                    if 0.3 < aspect_ratio < 5.0:  # 更宽松的长宽比
                        valid_contours.append(contour)
            
            if not valid_contours:
                return None
            
            # 智能评分系统
            best_contour = None
            best_score = -1
            
            for contour in valid_contours:
                area = cv2.contourArea(contour)
                x, y, w_rect, h_rect = cv2.boundingRect(contour)
                
                # 1. 面积得分（适中最好）
                area_score = 1.0 - abs(area - image_area * 0.2) / (image_area * 0.2)
                
                # 2. 位置得分（越居中越好）
                center_x = x + w_rect // 2
                center_y = y + h_rect // 2
                image_center_x = w // 2
                image_center_y = h // 2
                
                distance = np.sqrt((center_x - image_center_x)**2 + (center_y - image_center_y)**2)
                max_distance = np.sqrt((w//2)**2 + (h//2)**2)
                position_score = 1.0 - (distance / max_distance)
                
                # 3. 形状得分（越规则越好）
                perimeter = cv2.arcLength(contour, True)
                shape_score = 4 * np.pi * area / (perimeter * perimeter) if perimeter > 0 else 0
                
                # 4. 综合得分
                total_score = area_score * 0.4 + position_score * 0.4 + shape_score * 0.2
                
                if total_score > best_score:
                    best_score = total_score
                    best_contour = contour
            
            if best_contour is not None:
                # 创建最终掩码
                final_mask = np.zeros_like(mask)
                cv2.fillPoly(final_mask, [best_contour], 255)
                
                # 边缘优化
                kernel = np.ones((3, 3), np.uint8)
                final_mask = cv2.dilate(final_mask, kernel, iterations=1)
                
                return final_mask
            
            return None
            
        except Exception as e:
            print(f"智能轮廓筛选失败: {e}")
            return None
    
    def edge_based_bridge_detection(self, image):
        """基于边缘检测的桥梁识别"""
        try:
            # 转换为灰度图
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            # 高斯模糊减少噪点
            blurred = cv2.GaussianBlur(gray, (5, 5), 0)
            
            # Canny边缘检测
            edges = cv2.Canny(blurred, 50, 150)
            
            # 查找轮廓
            contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            if contours:
                # 找到最大的轮廓
                largest_contour = max(contours, key=cv2.contourArea)
                area = cv2.contourArea(largest_contour)
                
                if area > 1000:
                    # 创建掩码
                    mask = np.zeros_like(gray)
                    cv2.fillPoly(mask, [largest_contour], 255)
                    
                    # 形态学操作清理
                    kernel = np.ones((3, 3), np.uint8)
                    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
                    
                    return mask
            
            # 最后的备用方案：智能区域检测
            return self.smart_region_detection(image)
            
        except Exception as e:
            print(f"边缘检测桥梁识别失败: {e}")
            return self.smart_region_detection(image)
    
    def assess_image_quality(self, image):
        """评估图像质量，识别主要特征"""
        try:
            hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
            h, w = image.shape[:2]
            
            # 1. 检测天空区域
            # 天空：高亮度、低饱和度的蓝色
            sky_mask = cv2.inRange(hsv, np.array([100, 10, 180]), np.array([140, 60, 255]))
            sky_ratio = np.sum(sky_mask > 0) / (h * w)
            has_clear_sky = sky_ratio > 0.1  # 天空占比超过10%
            
            # 2. 检测植被区域
            # 植被：中等亮度、高饱和度的绿色
            vegetation_mask = cv2.inRange(hsv, np.array([35, 80, 40]), np.array([85, 255, 200]))
            vegetation_ratio = np.sum(vegetation_mask > 0) / (h * w)
            has_vegetation = vegetation_ratio > 0.05  # 植被占比超过5%
            
            # 3. 检测水体区域
            # 水体：低亮度、低饱和度的深色
            water_mask = cv2.inRange(hsv, np.array([0, 0, 0]), np.array([180, 100, 100]))
            water_ratio = np.sum(water_mask > 0) / (h * w)
            has_water = water_ratio > 0.1  # 水体占比超过10%
            
            # 4. 检测桥梁材质
            # 桥梁：浅灰到白色
            bridge_mask = cv2.inRange(hsv, np.array([0, 0, 100]), np.array([180, 60, 255]))
            bridge_ratio = np.sum(bridge_mask > 0) / (h * w)
            has_bridge_material = bridge_ratio > 0.05  # 桥梁材质占比超过5%
            
            return {
                'has_clear_sky': has_clear_sky,
                'has_vegetation': has_vegetation,
                'has_water': has_water,
                'has_bridge_material': has_bridge_material,
                'sky_ratio': sky_ratio,
                'vegetation_ratio': vegetation_ratio,
                'water_ratio': water_ratio,
                'bridge_ratio': bridge_ratio
            }
            
        except Exception as e:
            print(f"图像质量评估失败: {e}")
            return {
                'has_clear_sky': False,
                'has_vegetation': False,
                'has_water': False,
                'has_bridge_material': False,
                'sky_ratio': 0,
                'vegetation_ratio': 0,
                'water_ratio': 0,
                'bridge_ratio': 0
            }
    
    def sky_filter_strategy(self, image):
        """天空过滤策略（针对有清晰天空的图像）"""
        try:
            hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
            
            # 1. 桥梁材质检测（更严格的范围）
            lower_bridge = np.array([0, 0, 110])      # 进一步提高明度下限
            upper_bridge = np.array([180, 50, 255])   # 进一步降低饱和度上限
            
            bridge_mask = cv2.inRange(hsv, lower_bridge, upper_bridge)
            
            # 2. 多级天空过滤
            # 主天空：明显的蓝色
            sky_main = cv2.inRange(hsv, np.array([100, 10, 200]), np.array([140, 60, 255]))
            
            # 次天空：浅蓝色
            sky_secondary = cv2.inRange(hsv, np.array([100, 5, 180]), np.array([140, 80, 255]))
            
            # 合并天空掩码
            sky_mask = cv2.bitwise_or(sky_main, sky_secondary)
            
            # 3. 从桥梁掩码中减去天空
            bridge_mask = cv2.bitwise_and(bridge_mask, cv2.bitwise_not(sky_mask))
            
            # 4. 智能形态学清理
            bridge_mask = self.smart_morphology_cleanup(bridge_mask)
            
            # 5. 轮廓筛选和优化
            return self.contour_optimization(image, bridge_mask)
            
        except Exception as e:
            print(f"天空过滤策略失败: {e}")
            return self.simple_fallback_detection(image)
    
    def vegetation_filter_strategy(self, image):
        """植被过滤策略（针对有植被的图像）"""
        try:
            hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
            
            # 1. 桥梁材质检测
            lower_bridge = np.array([0, 0, 100])
            upper_bridge = np.array([180, 60, 255])
            bridge_mask = cv2.inRange(hsv, lower_bridge, upper_bridge)
            
            # 2. 植被过滤（多级）
            # 主要植被：绿色
            vegetation_main = cv2.inRange(hsv, np.array([35, 80, 40]), np.array([85, 255, 200]))
            
            # 次要植被：黄绿色
            vegetation_secondary = cv2.inRange(hsv, np.array([25, 60, 50]), np.array([95, 255, 220]))
            
            # 合并植被掩码
            vegetation_mask = cv2.bitwise_or(vegetation_main, vegetation_secondary)
            
            # 3. 从桥梁掩码中减去植被
            bridge_mask = cv2.bitwise_and(bridge_mask, cv2.bitwise_not(vegetation_mask))
            
            # 4. 智能形态学清理
            bridge_mask = self.smart_morphology_cleanup(bridge_mask)
            
            # 5. 轮廓筛选和优化
            return self.contour_optimization(image, bridge_mask)
            
        except Exception as e:
            print(f"植被过滤策略失败: {e}")
            return self.simple_fallback_detection(image)
    
    def water_filter_strategy(self, image):
        """水体过滤策略（针对有水的图像）"""
        try:
            hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
            
            # 1. 桥梁材质检测
            lower_bridge = np.array([0, 0, 100])
            upper_bridge = np.array([180, 60, 255])
            bridge_mask = cv2.inRange(hsv, lower_bridge, upper_bridge)
            
            # 2. 水体过滤
            water_mask = cv2.inRange(hsv, np.array([0, 0, 0]), np.array([180, 100, 120]))
            
            # 3. 从桥梁掩码中减去水体
            bridge_mask = cv2.bitwise_and(bridge_mask, cv2.bitwise_not(water_mask))
            
            # 4. 智能形态学清理
            bridge_mask = self.smart_morphology_cleanup(bridge_mask)
            
            # 5. 轮廓筛选和优化
            return self.contour_optimization(image, bridge_mask)
            
        except Exception as e:
            print(f"水体过滤策略失败: {e}")
            return self.simple_fallback_detection(image)
    
    def multi_feature_strategy(self, image):
        """多特征融合策略（针对复杂场景）"""
        try:
            # 1. 颜色特征
            color_mask = self.multi_color_analysis(image)
            
            # 2. 边缘特征
            edge_mask = self.edge_analysis(image)
            
            # 3. 纹理特征
            texture_mask = self.texture_analysis(image)
            
            # 4. 特征融合
            fused_mask = cv2.addWeighted(color_mask, 0.6, edge_mask, 0.3, 0)
            fused_mask = cv2.addWeighted(fused_mask, 0.8, texture_mask, 0.2, 0)
            
            # 5. 二值化
            _, fused_mask = cv2.threshold(fused_mask, 127, 255, cv2.THRESH_BINARY)
            
            # 6. 智能形态学清理
            fused_mask = self.smart_morphology_cleanup(fused_mask)
            
            # 7. 轮廓筛选和优化
            return self.contour_optimization(image, fused_mask)
            
        except Exception as e:
            print(f"多特征融合策略失败: {e}")
            return self.simple_fallback_detection(image)
    
    def smart_morphology_cleanup(self, mask):
        """智能形态学清理"""
        try:
            # 1. 分析掩码结构
            contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            if not contours:
                return mask
            
            # 2. 计算掩码特性
            total_area = np.sum(mask > 0)
            largest_contour = max(contours, key=cv2.contourArea)
            largest_area = cv2.contourArea(largest_contour)
            
            # 3. 根据结构选择形态学操作
            if largest_area / total_area > 0.8:
                # 主要是一个大区域：温和清理
                kernel_open = np.ones((3, 3), np.uint8)
                kernel_close = np.ones((7, 7), np.uint8)
            else:
                # 多个小区域：强力清理
                kernel_open = np.ones((5, 5), np.uint8)
                kernel_close = np.ones((9, 9), np.uint8)
            
            # 4. 形态学操作
            mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel_open)
            mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel_close)
            
            return mask
            
        except Exception as e:
            print(f"智能形态学清理失败: {e}")
            return mask
    
    def contour_optimization(self, image, mask):
        """轮廓优化和筛选"""
        try:
            contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            if not contours:
                return None
            
            h, w = image.shape[:2]
            image_area = h * w
            
            # 1. 面积过滤
            min_area = image_area * 0.005  # 0.5%
            max_area = image_area * 0.8    # 80%
            
            valid_contours = []
            for contour in contours:
                area = cv2.contourArea(contour)
                if min_area < area < max_area:
                    valid_contours.append(contour)
            
            if not valid_contours:
                return None
            
            # 2. 智能评分系统
            best_contour = None
            best_score = -1
            
            for contour in valid_contours:
                area = cv2.contourArea(contour)
                x, y, w_rect, h_rect = cv2.boundingRect(contour)
                
                # 面积得分
                area_score = 1.0 - abs(area - image_area * 0.15) / (image_area * 0.15)
                
                # 位置得分（越居中越好）
                center_x = x + w_rect // 2
                center_y = y + h_rect // 2
                image_center_x = w // 2
                image_center_y = h // 2
                
                distance = np.sqrt((center_x - image_center_x)**2 + (center_y - image_center_y)**2)
                max_distance = np.sqrt((w//2)**2 + (h//2)**2)
                position_score = 1.0 - (distance / max_distance)
                
                # 形状得分
                perimeter = cv2.arcLength(contour, True)
                shape_score = 4 * np.pi * area / (perimeter * perimeter) if perimeter > 0 else 0
                
                # 综合得分
                total_score = area_score * 0.4 + position_score * 0.4 + shape_score * 0.2
                
                if total_score > best_score:
                    best_score = total_score
                    best_contour = contour
            
            if best_contour is not None:
                # 创建最终掩码
                final_mask = np.zeros_like(mask)
                cv2.fillPoly(final_mask, [best_contour], 255)
                
                # 边缘优化
                kernel = np.ones((3, 3), np.uint8)
                final_mask = cv2.dilate(final_mask, kernel, iterations=1)
                
                return final_mask
            
            return None
            
        except Exception as e:
            print(f"轮廓优化失败: {e}")
            return None
    
    def simple_fallback_detection(self, image):
        """简单备用检测（基于灰度阈值）"""
        try:
            # 转换为灰度图
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            # 使用Otsu自动阈值
            _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            # 形态学清理
            kernel = np.ones((7, 7), np.uint8)
            binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)
            binary = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel)
            
            # 找到最大轮廓
            contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            if contours:
                largest_contour = max(contours, key=cv2.contourArea)
                area = cv2.contourArea(largest_contour)
                
                # 面积过滤
                h, w = image.shape[:2]
                if area > h * w * 0.005:  # 至少0.5%
                    mask = np.zeros_like(gray)
                    cv2.fillPoly(mask, [largest_contour], 255)
                    return mask
            
            # 最后的备用方案：中心矩形
            return self.smart_region_detection(image)
            
        except Exception as e:
            print(f"简单备用检测失败: {e}")
            return self.smart_region_detection(image)
    
    def smart_region_detection(self, image):
        """智能区域检测"""
        try:
            h, w = image.shape[:2]
            
            # 创建掩码
            mask = np.zeros((h, w), dtype=np.uint8)
            
            # 在图像中心区域寻找桥梁
            center_x, center_y = w // 2, h // 2
            search_width = int(w * 0.6)   # 搜索宽度为图像宽度的60%
            search_height = int(h * 0.4)  # 搜索高度为图像高度的40%
            
            # 在搜索区域内寻找可能的桥梁
            start_x = max(0, center_x - search_width // 2)
            end_x = min(w, center_x + search_width // 2)
            start_y = max(0, center_y - search_height // 2)
            end_y = min(h, center_y + search_height // 2)
            
            # 在搜索区域内填充
            cv2.rectangle(mask, (start_x, start_y), (end_x, end_y), 255, -1)
            
            return mask
            
        except Exception as e:
            print(f"智能区域检测失败: {e}")
            # 返回一个简单的默认掩码
            h, w = image.shape[:2]
            default_mask = np.zeros((h, w), dtype=np.uint8)
            cv2.rectangle(default_mask, (w//4, h//4), (3*w//4, 3*h//4), 255, -1)
            return default_mask
    

    
    def extract_hsv_data(self, image, mask):
        """提取HSV数据"""
        try:
            hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
            masked_hsv = cv2.bitwise_and(hsv, hsv, mask=mask)
            
            non_zero_pixels = cv2.findNonZero(mask)
            if non_zero_pixels is None:
                return None
            
            hsv_values = []
            for point in non_zero_pixels:
                x, y = point[0]
                h, s, v = masked_hsv[y, x]
                hsv_values.append([h, s, v])
            
            return np.array(hsv_values)
            
        except Exception as e:
            print(f"HSV数据提取出错: {e}")
            return None
    
    def display_current_image(self):
        """显示当前图像"""
        if 0 <= self.current_image_index < len(self.image_paths):
            # 尝试读取原始图像并显示
            image_path = self.image_paths[self.current_image_index]
            
            # 读取图像
            image = self.read_image_safely(image_path)
            if image is not None:
                self.display_image_on_label(self.original_image_label, image)
            else:
                self.original_image_label.config(image='')
                self.original_image_label.image = None

            # 显示对应的掩码（如果已生成）
            if 0 <= self.current_image_index < len(self.bridge_masks) and self.bridge_masks[self.current_image_index] is not None:
                mask = self.bridge_masks[self.current_image_index]
                # 将掩码转换为高质量可视化图像
                if len(mask.shape) == 2:  # 单通道掩码
                    # 创建高质量彩色掩码显示
                    mask_colored = self.create_high_quality_mask(mask)
                    self.display_image_on_label(self.mask_image_label, mask_colored)
                else:
                    self.display_image_on_label(self.mask_image_label, mask)
            else:
                # 显示默认的"无掩码"提示
                self.mask_image_label.config(text="等待分析...", font=('Arial', 16), fg='gray', bg='white')
                self.mask_image_label.image = None
        else:
            self.original_image_label.config(image='')
            self.mask_image_label.config(image='')
            self.mask_image_label.config(image='')
            self.mask_image_label.image = None
        self.update_image_count_label()
    
    def display_image_on_label(self, label, img_cv):
        """在标签上显示图像"""
        if img_cv is None:
            label.config(image='')
            label.image = None
            return

        try:
            # 确保图像是3通道
            if len(img_cv.shape) == 2:  # 单通道图像
                img_cv = cv2.cvtColor(img_cv, cv2.COLOR_GRAY2BGR)
            
            # 调整图像大小以适应标签
            h, w = img_cv.shape[:2]
            label_width = 400  # 固定宽度
            label_height = 300  # 固定高度

            if w > label_width or h > label_height:
                scale_w = label_width / w
                scale_h = label_height / h
                scale = min(scale_w, scale_h)
                new_w, new_h = int(w * scale), int(h * scale)
                img_cv = cv2.resize(img_cv, (new_w, new_h), interpolation=cv2.INTER_AREA)

            # 转换为RGB并创建PhotoImage
            img_rgb = cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB)
            img_pil = Image.fromarray(img_rgb)
            img_tk = ImageTk.PhotoImage(image=img_pil)
            
            # 更新标签
            label.config(image=img_tk)
            label.image = img_tk  # 保持引用
            
        except Exception as e:
            print(f"图像显示错误: {e}")
            label.config(image='')
            label.image = None
    
    def update_image_count_label(self):
        """更新图像计数标签"""
        total = len(self.image_paths)
        current = self.current_image_index + 1 if total > 0 else 0
        self.image_count_label.config(text=f"{current}/{total}")

    def show_previous_image(self):
        """显示上一张图像"""
        if self.current_image_index > 0:
            self.current_image_index -= 1
            self.display_current_image()

    def show_next_image(self):
        """显示下一张图像"""
        if self.current_image_index < len(self.image_paths) - 1:
            self.current_image_index += 1
            self.display_current_image()

    def show_hsv_histograms(self):
        """显示HSV直方图"""
        if not self.hsv_data:
            messagebox.showwarning("警告", "请先分析桥梁颜色！")
            return
        
        all_hsv = np.vstack(self.hsv_data)
        
        fig, axes = plt.subplots(3, 1, figsize=(10, 8))
        fig.suptitle('Bridge HSV Color Distribution Histograms', fontsize=16, fontweight='bold')
        
        # H直方图
        axes[0].hist(all_hsv[:, 0], bins=36, range=(0, 180), alpha=0.7, color='blue', edgecolor='black')
        axes[0].set_title('Hue (H) Histogram', fontsize=12, fontweight='bold')
        axes[0].set_xlabel('Hue Value')
        axes[0].set_ylabel('Pixel Count')
        axes[0].grid(True, alpha=0.3)
        
        # S直方图
        axes[1].hist(all_hsv[:, 1], bins=50, range=(0, 255), alpha=0.7, color='green', edgecolor='black')
        axes[1].set_title('Saturation (S) Histogram', fontsize=12, fontweight='bold')
        axes[1].set_xlabel('Saturation Value')
        axes[1].set_ylabel('Pixel Count')
        axes[1].grid(True, alpha=0.3)
        
        # V直方图
        axes[2].hist(all_hsv[:, 2], bins=50, range=(0, 255), alpha=0.7, color='red', edgecolor='black')
        axes[2].set_title('Value (V) Histogram', fontsize=12, fontweight='bold')
        axes[2].set_xlabel('Value')
        axes[2].set_ylabel('Pixel Count')
        axes[2].grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.show()
    
    def show_3d_scatter(self):
        """显示3D散点图"""
        if not self.hsv_data:
            messagebox.showwarning("警告", "请先分析桥梁颜色！")
            return
        
        all_hsv = np.vstack(self.hsv_data)
        
        if len(all_hsv) > 10000:
            indices = np.random.choice(len(all_hsv), 10000, replace=False)
            sampled_hsv = all_hsv[indices]
        else:
            sampled_hsv = all_hsv
        
        fig = plt.figure(figsize=(12, 8))
        ax = fig.add_subplot(111, projection='3d')
        
        scatter = ax.scatter(sampled_hsv[:, 0], sampled_hsv[:, 1], sampled_hsv[:, 2], 
                           c=sampled_hsv[:, 0], cmap='hsv', alpha=0.6, s=1)
        
        ax.set_xlabel('Hue (H)', fontsize=12, fontweight='bold')
        ax.set_ylabel('Saturation (S)', fontsize=12, fontweight='bold')
        ax.set_zlabel('Value (V)', fontsize=12, fontweight='bold')
        ax.set_title('Bridge HSV Color Distribution 3D Scatter Plot', fontsize=16, fontweight='bold')
        
        cbar = plt.colorbar(scatter, ax=ax, shrink=0.8, aspect=20)
        cbar.set_label('Hue Value', fontsize=10)
        
        ax.set_xlim(0, 180)
        ax.set_ylim(0, 255)
        ax.set_zlim(0, 255)
        
        plt.tight_layout()
        plt.show()
    
    def clear_data(self):
        """清除数据"""
        self.image_paths = []
        self.images = []
        self.bridge_masks = []
        self.hsv_data = []
        self.color_analysis_results = []
        self.current_image_index = 0
        
        self.original_image_label.config(image='')
        self.mask_image_label.config(image='')
        self.status_label.config(text="状态: 数据已清除，等待选择图像...")
        self.progress_bar['value'] = 0
        self.update_image_count_label()
        
        # 清除颜色分析显示
        for widget in self.color_results_frame.winfo_children():
            widget.destroy()
        
        messagebox.showinfo("清除", "所有分析数据已清除。")
    
    
    def extract_segment_and_number_from_filename(self, filename):
        """从文件名中提取分段字母和数字序号用于排序"""
        import re
        # 移除文件扩展名
        name_without_ext = os.path.splitext(filename)[0]
        
        # 提取分段字母（A、B、C、D、E、F）
        segment_match = re.search(r'^([A-F])', name_without_ext.upper())
        segment = segment_match.group(1) if segment_match else 'A'
        
        # 提取数字序号
        numbers = re.findall(r'\d+', name_without_ext)
        number = int(numbers[0]) if numbers else 0
        
        return segment, number
    
    def export_hsv_data(self):
        """导出简化的HSV数据（F-H, F-S, F-V）"""
        if not self.hsv_data:
            messagebox.showwarning("警告", "请先分析桥梁颜色！")
            return
        
        try:
            # 选择保存路径
            file_path = filedialog.asksaveasfilename(
                title="保存HSV数据文件",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            
            if not file_path:
                return
            
            # 导入openpyxl库
            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                messagebox.showerror("错误", "需要安装openpyxl库！\n\n请运行：pip install openpyxl")
                return
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "桥梁HSV数据"
            
            # 设置标题行（按照您的表格格式）
            headers = ["序号", "纬度", "经度", "图像", "F- H", "F- S", "F- V"]
            
            # 设置标题样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 创建排序索引：按照分段字母和数字序号排序
            sorted_indices = []
            for i, image_path in enumerate(self.image_paths):
                filename = os.path.basename(image_path)
                segment, number = self.extract_segment_and_number_from_filename(filename)
                sorted_indices.append((segment, number, i, filename))
            
            # 按照分段字母和数字序号排序
            # 首先按分段字母排序（A、B、C、D、E、F），然后按数字序号排序
            sorted_indices.sort(key=lambda x: (x[0], x[1]))
            
            # 填充数据（按照排序后的顺序）
            row = 2
            for _, _, original_index, filename in sorted_indices:
                if original_index < len(self.hsv_data):
                    hsv_data = self.hsv_data[original_index]
                    if hsv_data is None or len(hsv_data) == 0:
                        continue
                    
                    # 序号（按照排序后的顺序）
                    ws.cell(row=row, column=1, value=f"F{row-1}")
                    
                    # 纬度（这里需要您手动填入，或者从文件名提取）
                    ws.cell(row=row, column=2, value="")  # 留空，需要手动填入
                    
                    # 经度（这里需要您手动填入，或者从文件名提取）
                    ws.cell(row=row, column=3, value="")  # 留空，需要手动填入
                    
                    # 图像名称
                    ws.cell(row=row, column=4, value=filename)
                    
                    # F- H (平均色调值)
                    avg_h = np.mean(hsv_data[:, 0])
                    ws.cell(row=row, column=5, value=round(avg_h, 2))
                    
                    # F- S (平均饱和度值)
                    avg_s = np.mean(hsv_data[:, 1])
                    ws.cell(row=row, column=6, value=round(avg_s, 2))
                    
                    # F- V (平均明度值)
                    avg_v = np.mean(hsv_data[:, 2])
                    ws.cell(row=row, column=7, value=round(avg_v, 2))
                    
                    row += 1
            
            # 调整列宽
            column_widths = [8, 12, 12, 20, 10, 10, 10]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # 保存文件
            wb.save(file_path)
            
            messagebox.showinfo("成功", f"HSV数据已成功导出到：\n{file_path}\n\n文件包含 {len(self.hsv_data)} 张图像的分析结果。\n\n已按照分段顺序排序：\n- 首先按分段字母排序（A、B、C、D、E、F）\n- 每个分段内按数字序号排序\n\n注意：纬度和经度需要您手动填入！")
            
        except Exception as e:
            messagebox.showerror("错误", f"导出HSV数据时出错：\n{str(e)}")
            print(f"HSV数据导出错误: {e}")
    
    def show_segment_histograms(self):
        """按段显示HSV直方图（A、B、C、D、E、F段）"""
        if not self.hsv_data or not self.image_paths:
            messagebox.showwarning("警告", "请先分析桥梁颜色！")
            return
        
        try:
            # 按段分组数据
            segment_data = self.group_data_by_segment()
            
            if not segment_data:
                messagebox.showwarning("警告", "无法识别图像段位！请确保图像文件名以A、B、C、D、E、F开头。")
                return
            
            # 创建子图
            fig, axes = plt.subplots(2, 3, figsize=(18, 12))
            fig.suptitle('Bridge HSV Color Distribution - Segment Grouped Histograms', fontsize=16, fontweight='bold')
            
            # 定义段位顺序
            segments = ['A', 'B', 'C', 'D', 'E', 'F']
            
            for i, segment in enumerate(segments):
                row = i // 3
                col = i % 3
                ax = axes[row, col]
                
                if segment in segment_data and len(segment_data[segment]) > 0:
                    # 获取该段的所有HSV数据
                    segment_hsv = np.vstack(segment_data[segment])
                    
                    # 绘制H、S、V三个直方图
                    ax.hist(segment_hsv[:, 0], bins=18, range=(0, 180), alpha=0.7, 
                           label='Hue (H)', color='red', edgecolor='black')
                    ax.hist(segment_hsv[:, 1], bins=25, range=(0, 255), alpha=0.7, 
                           label='Saturation (S)', color='green', edgecolor='black')
                    ax.hist(segment_hsv[:, 2], bins=25, range=(0, 255), alpha=0.7, 
                           label='Value (V)', color='blue', edgecolor='black')
                    
                    ax.set_title(f'{segment} Segment HSV Distribution', fontsize=12, fontweight='bold')
                    ax.set_xlabel('HSV Value')
                    ax.set_ylabel('Pixel Count')
                    ax.legend()
                    ax.grid(True, alpha=0.3)
                    
                    # 添加统计信息
                    avg_h = np.mean(segment_hsv[:, 0])
                    avg_s = np.mean(segment_hsv[:, 1])
                    avg_v = np.mean(segment_hsv[:, 2])
                    ax.text(0.02, 0.98, f'Avg H: {avg_h:.1f}\nAvg S: {avg_s:.1f}\nAvg V: {avg_v:.1f}', 
                           transform=ax.transAxes, verticalalignment='top',
                           bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
                else:
                    ax.set_title(f'{segment} Segment (No Data)', fontsize=12, fontweight='bold')
                    ax.text(0.5, 0.5, 'No Data', transform=ax.transAxes, 
                           horizontalalignment='center', verticalalignment='center',
                           fontsize=14, color='gray')
                    ax.set_xlabel('HSV Value')
                    ax.set_ylabel('Pixel Count')
            
            plt.tight_layout()
            plt.show()
            
        except Exception as e:
            messagebox.showerror("错误", f"显示按段直方图时出错：\n{str(e)}")
            print(f"按段直方图错误: {e}")
    
    def group_data_by_segment(self):
        """按段分组数据（A、B、C、D、E、F段）"""
        segment_data = {'A': [], 'B': [], 'C': [], 'D': [], 'E': [], 'F': []}
        
        for i, (image_path, hsv_data) in enumerate(zip(self.image_paths, self.hsv_data)):
            if hsv_data is None or len(hsv_data) == 0:
                continue
            
            # 从文件名提取段位
            filename = os.path.basename(image_path)
            segment = self.extract_segment_from_filename(filename)
            
            if segment in segment_data:
                segment_data[segment].append(hsv_data)
        
        return segment_data
    
    def extract_segment_from_filename(self, filename):
        """从文件名提取段位（A、B、C、D、E、F）"""
        # 移除文件扩展名
        name_without_ext = os.path.splitext(filename)[0]
        
        # 查找段位标识
        for char in name_without_ext:
            if char.upper() in ['A', 'B', 'C', 'D', 'E', 'F']:
                return char.upper()
        
        # 如果没有找到明确的段位标识，尝试从文件名模式推断
        if name_without_ext.startswith('A'):
            return 'A'
        elif name_without_ext.startswith('B'):
            return 'B'
        elif name_without_ext.startswith('C'):
            return 'C'
        elif name_without_ext.startswith('D'):
            return 'D'
        elif name_without_ext.startswith('E'):
            return 'E'
        elif name_without_ext.startswith('F'):
            return 'F'
        
        # 默认返回A段
        return 'A'
    
    def create_high_quality_mask(self, mask):
        """创建黑白掩码显示"""
        try:
            # 创建黑白掩码（更清晰直观）
            # 桥梁区域：白色 (255, 255, 255)
            # 背景区域：黑色 (0, 0, 0)
            mask_bw = np.zeros((mask.shape[0], mask.shape[1], 3), dtype=np.uint8)
            mask_bw[mask > 0] = [255, 255, 255]  # 桥梁区域显示为白色
            return mask_bw
            
        except Exception as e:
            print(f"掩码创建失败: {e}")
            # 备用方案：简单黑白掩码
            mask_bw = np.zeros((mask.shape[0], mask.shape[1], 3), dtype=np.uint8)
            mask_bw[mask > 0] = [255, 255, 255]
            return mask_bw
    
    def run(self):
        """运行程序"""
        self.root.mainloop()

if __name__ == "__main__":
    try:
        import cv2
        import numpy as np
        import matplotlib.pyplot as plt
        from PIL import Image, ImageTk
    except ImportError as e:
        print(f"缺少必要的依赖库: {e}")
        print("请安装以下库:")
        print("pip install opencv-python numpy matplotlib pillow")
        exit(1)
    
    app = 桥梁颜色分析器()
    app.run()
